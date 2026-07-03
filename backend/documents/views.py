"""
REST API views for documents app
Supports both local and Supabase Storage
"""
import logging
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.utils import timezone
from django.db import transaction
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
import uuid
import os

from .models import AadhaarDocument, DocumentMetadata
from .serializers import (
    AadhaarDocumentSerializer,
    DocumentMetadataSerializer,
    DocumentUploadSerializer,
    BatchProcessSerializer,
)
from .preprocessing import ImagePreprocessor
from .gemini_service import GeminiService
from .storage_service import get_storage_service

logger = logging.getLogger(__name__)


class AadhaarDocumentViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Aadhaar document operations
    
    Provides CRUD operations and additional actions for document processing
    Documents are filtered by the authenticated user.
    """
    serializer_class = AadhaarDocumentSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)
    
    def get_permissions(self):
        """
        Set permissions based on action:
        - list, retrieve: Allow any (public can view counts)
        - create, update, delete, upload: Require authentication
        """
        if self.action in ['upload', 'create', 'update', 'partial_update', 'destroy', 
                           'analyze', 'batch_analyze', 'batch_delete']:
            return [IsAuthenticated()]
        return [AllowAny()]
    
    def get_queryset(self):
        """
        Filter documents by authenticated user.
        Unauthenticated users see no documents.
        Uses select_related for optimized queries.
        """
        if self.request.user.is_authenticated:
            return (
                AadhaarDocument.objects
                .filter(user=self.request.user)
                .select_related('metadata')  # Optimize: fetch metadata in single query
                .order_by('-uploaded_at')  # Most recent first
            )
        return AadhaarDocument.objects.none()
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def upload(self, request):
        """
        Upload single or multiple Aadhaar documents
        
        Accepts:
        - files: List of image files (required)
        - batch_id: Optional batch identifier (auto-generated if not provided)
        - auto_analyze: Whether to automatically analyze with Gemini (default: True)
        
        Returns:
        - List of created document objects
        
        Storage:
        - Uses Supabase Storage when USE_SUPABASE_STORAGE=true
        - Uses local filesystem when USE_SUPABASE_STORAGE=false
        """
        from django.conf import settings as django_settings
        from io import BytesIO
        
        # Handle both single file and multiple files
        files = request.FILES.getlist('files')
        if not files:
            # Try single file upload
            single_file = request.FILES.get('file')
            if single_file:
                files = [single_file]
            else:
                return Response(
                    {'error': 'No files provided. Use "files" or "file" field.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        batch_id = request.data.get('batch_id')
        if not batch_id and len(files) > 1:
            # Auto-generate batch ID for multiple files
            batch_id = f"batch_{uuid.uuid4().hex[:12]}"
        
        auto_analyze = request.data.get('auto_analyze', 'true').lower() == 'true'
        
        # Get storage service (handles both local and Supabase based on settings)
        storage_service = get_storage_service()
        use_supabase = storage_service.use_supabase
        
        created_documents = []
        failed_files = []
        
        for idx, file in enumerate(files):
            try:
                # Each file gets its own transaction to avoid breaking the entire batch
                with transaction.atomic():
                    # Read file bytes once for reuse
                    file_bytes = file.read()
                    file.seek(0)  # Reset for potential Django ImageField use
                    
                    if use_supabase:
                        # === SUPABASE STORAGE MODE ===
                        # Create document record WITHOUT local file (will set paths later)
                        document = AadhaarDocument.objects.create(
                            user=request.user,
                            file_name=file.name,
                            file_size=file.size,
                            status='uploaded',
                            batch_id=batch_id,
                            batch_position=idx if batch_id else None,
                            storage_type='supabase',
                        )
                        
                        # Upload original file to Supabase
                        original_result = storage_service.upload_file(
                            file_data=file_bytes,
                            user_id=request.user.id,
                            document_id=document.id,
                            filename=file.name,
                            folder='raw',
                            content_type=file.content_type
                        )
                        document.supabase_original_path = original_result['path']
                        document.save()
                        
                        # Preprocess the image from bytes
                        preprocessor = ImagePreprocessor(BytesIO(file_bytes))
                        preprocess_result = preprocessor.process_all()
                        
                        # Get preprocessed image bytes and upload to Supabase
                        processed_bytes, processed_content_type = preprocessor.to_bytes()
                        processed_filename = f"proc_{document.id}_{file.name}"
                        processed_result = storage_service.upload_file(
                            file_data=processed_bytes,
                            user_id=request.user.id,
                            document_id=document.id,
                            filename=processed_filename,
                            folder='processed',
                            content_type=processed_content_type
                        )
                        document.supabase_processed_path = processed_result['path']
                        
                        # Create thumbnail and upload to Supabase
                        thumb_preprocessor = ImagePreprocessor(BytesIO(file_bytes))
                        thumb = thumb_preprocessor.create_thumbnail()
                        thumb_bytes_io = BytesIO()
                        thumb.save(thumb_bytes_io, format='JPEG', quality=85)
                        thumb_bytes_io.seek(0)  # Reset pointer to beginning
                        thumb_bytes = thumb_bytes_io.getvalue()
                        
                        logger.info(f"Thumbnail size: {len(thumb_bytes)} bytes for document {document.id}")
                        
                        thumb_filename = f"thumb_{document.id}_{file.name}"
                        thumb_result = storage_service.upload_file(
                            file_data=thumb_bytes,
                            user_id=request.user.id,
                            document_id=document.id,
                            filename=thumb_filename,
                            folder='thumbnails',
                            content_type='image/jpeg'
                        )
                        document.supabase_thumbnail_path = thumb_result['path']
                        logger.info(f"Thumbnail uploaded to: {thumb_result['path']}")
                        
                        logger.info(f"Uploaded document {document.id} to Supabase: {original_result['path']}")
                        
                    else:
                        # === LOCAL STORAGE MODE ===
                        # Create document record with local file (Django ImageField)
                        document = AadhaarDocument.objects.create(
                            user=request.user,
                            original_file=file,
                            file_name=file.name,
                            file_size=file.size,
                            status='uploaded',
                            batch_id=batch_id,
                            batch_position=idx if batch_id else None,
                            storage_type='local',
                        )
                        
                        # Preprocess the image from local file path
                        preprocessor = ImagePreprocessor(document.original_file.path)
                        preprocess_result = preprocessor.process_all()
                        
                        # Save preprocessed image locally
                        preprocessed_filename = f"proc_{document.id}_{file.name}"
                        preprocessed_path = os.path.join('media', 'processed', preprocessed_filename)
                        preprocessor.save_to_file(preprocessed_path)
                        document.preprocessed_file = f"processed/{preprocessed_filename}"
                        
                        # Save thumbnail locally
                        thumb_preprocessor = ImagePreprocessor(document.original_file.path)
                        thumb = thumb_preprocessor.create_thumbnail()
                        thumb_filename = f"thumb_{document.id}_{file.name}"
                        thumb_path = os.path.join('media', 'thumbnails', thumb_filename)
                        os.makedirs(os.path.dirname(thumb_path), exist_ok=True)
                        thumb.save(thumb_path, quality=85)
                        document.thumbnail = f"thumbnails/{thumb_filename}"
                        
                        logger.info(f"Stored document {document.id} locally")
                    
                    document.status = 'processing'
                    document.save()
                    
                    # Create metadata record with quality report
                    metadata = DocumentMetadata.objects.create(
                        document=document,
                        quality_issues=preprocess_result['quality_report']['issues']
                    )
                    
                    # Run Gemini analysis if requested
                    if auto_analyze:
                        self._analyze_document(document, metadata)
                    
                    document.status = 'completed'
                    document.processed_at = timezone.now()
                    document.save()
                    
                    created_documents.append(document)
                    
            except Exception as e:
                # Log the error and continue with next file
                logger.error(f"Failed to process {file.name}: {str(e)}")
                failed_files.append({'file_name': file.name, 'error': str(e)})
                
                # Try to mark document as failed if it was created (outside the failed transaction)
                try:
                    if 'document' in locals() and document.id:
                        AadhaarDocument.objects.filter(id=document.id).update(
                            status='failed',
                            error_message=str(e)[:500]
                        )
                except Exception:
                    pass  # Ignore errors when trying to mark as failed
        
        # If no documents were created successfully, return error
        if not created_documents:
            return Response(
                {'error': f'Failed to process all files. Errors: {failed_files}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        # Serialize and return created documents
        serializer = self.get_serializer(created_documents, many=True, context={'request': request})
        
        response_data = {
            'success': True,
            'count': len(created_documents),
            'batch_id': batch_id,
            'storage_type': 'supabase' if use_supabase else 'local',
            'documents': serializer.data
        }
        
        # Include failed files info if any
        if failed_files:
            response_data['failed'] = failed_files
        
        return Response(response_data, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'])
    def analyze(self, request, pk=None):
        """
        Trigger Gemini analysis for a specific document
        
        Returns:
        - Updated document with analysis results
        """
        import traceback
        
        document = self.get_object()
        logger.info(f"Analyze request for document {document.id}, storage_type={document.storage_type}")
        
        if document.status == 'processing':
            return Response(
                {'error': 'Document is already being processed'},
                status=status.HTTP_409_CONFLICT
            )
        
        try:
            document.status = 'processing'
            document.save()
            
            # Get or create metadata
            metadata, created = DocumentMetadata.objects.get_or_create(document=document)
            
            # Run Gemini analysis
            self._analyze_document(document, metadata)
            
            document.status = 'completed'
            document.processed_at = timezone.now()
            document.save()
            
            serializer = self.get_serializer(document, context={'request': request})
            return Response(serializer.data)
            
        except Exception as e:
            # Log full traceback for debugging
            logger.error(f"Analyze failed for document {document.id}: {str(e)}")
            logger.error(traceback.format_exc())
            print(f"ERROR: {str(e)}")
            print(traceback.format_exc())
            
            document.status = 'failed'
            document.error_message = str(e)
            document.save()
            
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'])
    def batch_analyze(self, request):
        """
        Analyze multiple documents in batch
        
        Accepts:
        - document_ids: List of document IDs to analyze
        
        Returns:
        - Summary of batch processing results
        """
        serializer = BatchProcessSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        document_ids = serializer.validated_data['document_ids']
        # Filter by user to ensure user can only analyze their own documents
        documents = AadhaarDocument.objects.filter(id__in=document_ids, user=request.user)
        
        if documents.count() != len(document_ids):
            return Response(
                {'error': 'Some document IDs not found or access denied'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        results = {
            'total': len(document_ids),
            'successful': 0,
            'failed': 0,
            'details': []
        }
        
        for document in documents:
            try:
                document.status = 'processing'
                document.save()
                
                metadata, created = DocumentMetadata.objects.get_or_create(document=document)
                self._analyze_document(document, metadata)
                
                document.status = 'completed'
                document.processed_at = timezone.now()
                document.save()
                
                results['successful'] += 1
                results['details'].append({
                    'id': document.id,
                    'status': 'success',
                    'file_name': document.file_name
                })
                
            except Exception as e:
                document.status = 'failed'
                document.error_message = str(e)
                document.save()
                
                results['failed'] += 1
                results['details'].append({
                    'id': document.id,
                    'status': 'failed',
                    'file_name': document.file_name,
                    'error': str(e)
                })
        
        return Response(results)
    
    @action(detail=False, methods=['post'])
    def batch_delete(self, request):
        """
        Delete multiple documents in batch
        
        Accepts:
        - document_ids: List of document IDs to delete
        
        Returns:
        - Summary of batch delete results
        """
        serializer = BatchProcessSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        document_ids = serializer.validated_data['document_ids']
        # Filter by user to ensure user can only delete their own documents
        documents = AadhaarDocument.objects.filter(id__in=document_ids, user=request.user)
        
        found_count = documents.count()
        if found_count == 0:
            return Response(
                {'error': 'No documents found with the provided IDs or access denied'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        results = {
            'total': len(document_ids),
            'deleted': 0,
            'not_found': len(document_ids) - found_count,
            'details': []
        }
        
        # Get storage service for potential Supabase cleanup
        storage_service = get_storage_service()
        
        for document in documents:
            try:
                doc_info = {
                    'id': document.id,
                    'file_name': document.file_name,
                    'status': 'deleted'
                }
                
                # Delete associated files based on storage type
                if document.storage_type == 'supabase':
                    # Delete from Supabase Storage
                    if document.supabase_original_path:
                        storage_service.delete_file(document.supabase_original_path)
                    if document.supabase_processed_path:
                        storage_service.delete_file(document.supabase_processed_path)
                    if document.supabase_thumbnail_path:
                        storage_service.delete_file(document.supabase_thumbnail_path)
                    logger.info(f"Deleted document {document.id} from Supabase Storage")
                else:
                    # Delete from local storage via Django ImageField
                    if document.original_file:
                        document.original_file.delete(save=False)
                    if document.preprocessed_file:
                        document.preprocessed_file.delete(save=False)
                    if document.thumbnail:
                        document.thumbnail.delete(save=False)
                    logger.info(f"Deleted document {document.id} from local storage")
                
                # Delete the document record
                document.delete()
                
                results['deleted'] += 1
                results['details'].append(doc_info)
                
            except Exception as e:
                results['details'].append({
                    'id': document.id,
                    'file_name': document.file_name,
                    'status': 'failed',
                    'error': str(e)
                })
        
        return Response(results)
    
    @action(detail=False, methods=['get'])
    def batches(self, request):
        """
        List all batch IDs with document counts
        
        Returns:
        - List of batches with metadata
        """
        from django.db.models import Count
        
        # Filter by user
        queryset = self.get_queryset()
        batches = (
            queryset
            .exclude(batch_id__isnull=True)
            .exclude(batch_id='')
            .values('batch_id')
            .annotate(
                document_count=Count('id'),
            )
            .order_by('-batch_id')
        )
        
        return Response(list(batches))
    
    @action(detail=False, methods=['get'])
    def batch_documents(self, request):
        """
        Get all documents for a specific batch
        
        Query params:
        - batch_id: The batch identifier
        
        Returns:
        - List of documents in the batch
        """
        batch_id = request.query_params.get('batch_id')
        
        if not batch_id:
            return Response(
                {'error': 'batch_id parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Filter by user
        documents = self.get_queryset().filter(batch_id=batch_id).order_by('batch_position')
        serializer = self.get_serializer(documents, many=True, context={'request': request})
        
        return Response({
            'batch_id': batch_id,
            'count': documents.count(),
            'documents': serializer.data
        })
    
    @action(detail=False, methods=['get'])
    def verification_results(self, request):
        """
        Get verification results for completed documents with statistics
        
        Returns:
        - List of completed documents with verification status
        - Summary statistics (total, accepted, rejected)
        """
        from django.db.models import Q, Count
        
        # Filter completed documents with metadata (filtered by user via get_queryset)
        completed_docs = (
            self.get_queryset()
            .filter(status='completed')
            .filter(Q(metadata__isnull=False))
            .select_related('metadata')
            .order_by('-uploaded_at')
        )
        
        # Calculate statistics based on is_authentic field
        stats = {
            'total': completed_docs.count(),
            'accepted': completed_docs.filter(metadata__is_authentic=True).count(),
            'rejected': completed_docs.filter(metadata__is_authentic=False).count(),
        }
        
        # Serialize documents
        serializer = self.get_serializer(completed_docs, many=True, context={'request': request})
        
        return Response({
            'stats': stats,
            'documents': serializer.data,
            'count': completed_docs.count()
        })
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """
        Export verification results to Excel file
        
        Returns:
        - Excel file download with verification results
        """
        from django.db.models import Q
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        import datetime
        
        # Get completed documents with metadata (filtered by user via get_queryset)
        completed_docs = (
            self.get_queryset()
            .filter(status='completed')
            .filter(Q(metadata__isnull=False))
            .select_related('metadata')
            .order_by('-uploaded_at')
        )
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Verification Results"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers with all Aadhaar details and fraud detection data
        headers = [
            "Sr.No", "Image No.", "Document Type", "Status", "Final Remarks",
            "Aadhaar Number", "Name", "Date of Birth", "Gender", "Address",
            "Confidence Score", "Is Authentic", "Risk Score", "Risk Level",
            "YOLO Detections", "CV Analysis Results", "Fraud Indicators", "Quality Issues",
            "Upload Date", "Analyzed At"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows
        for row, doc in enumerate(completed_docs, 2):
            metadata = doc.metadata
            
            # Get verification status
            if metadata.is_authentic is True:
                status = "Accepted"
            elif metadata.is_authentic is False:
                status = "Rejected"
            else:
                status = "Pending"
            
            # Get document type
            doc_type = "Aadhaar" if metadata.aadhaar_number else "Non-Aadhaar"
            
            # Get final remarks
            if metadata.is_authentic is True:
                remarks = "All matched"
            elif metadata.is_authentic is False:
                if metadata.fraud_indicators:
                    remarks = metadata.fraud_indicators[0]
                elif metadata.quality_issues:
                    remarks = metadata.quality_issues[0]
                else:
                    remarks = "Verification failed"
            else:
                remarks = "Analysis pending"
            
            # Format fraud indicators and quality issues
            fraud_indicators_str = "; ".join(metadata.fraud_indicators) if metadata.fraud_indicators else "None"
            quality_issues_str = "; ".join(metadata.quality_issues) if metadata.quality_issues else "None"
            
            # Get fraud detection data
            fraud_detection = metadata.fraud_detection or {}
            risk_score = fraud_detection.get('risk_score', 0)
            risk_level = fraud_detection.get('risk_level', 'N/A')
            
            # Format YOLO detections
            yolo_detections = fraud_detection.get('yolo_detections', [])
            yolo_str = "; ".join([f"{det.get('class_name', det.get('class', 'Unknown'))} ({det.get('confidence', 0):.2f})" for det in yolo_detections]) if yolo_detections else "None"
            
            # Format CV Analysis results
            cv_analysis = fraud_detection.get('cv_analysis', {})
            cv_results = []
            if cv_analysis.get('compression_artifacts'):
                cv_results.append(f"Compression: {cv_analysis['compression_artifacts'].get('score', 0):.1f}")
            if cv_analysis.get('copy_paste_detection'):
                cv_results.append(f"Copy-Paste: {cv_analysis['copy_paste_detection'].get('score', 0):.1f}")
            if cv_analysis.get('noise_analysis'):
                cv_results.append(f"Noise: {cv_analysis['noise_analysis'].get('variance', 0):.1f}")
            if cv_analysis.get('edge_analysis'):
                cv_results.append(f"Edge: {cv_analysis['edge_analysis'].get('score', 0):.1f}")
            cv_str = "; ".join(cv_results) if cv_results else "None"
            
            # Write row data with all details
            ws.cell(row=row, column=1, value=row-1)  # Sr.No
            ws.cell(row=row, column=2, value=doc.file_name)  # Image No.
            ws.cell(row=row, column=3, value=doc_type)  # Document Type
            ws.cell(row=row, column=4, value=status)  # Status
            ws.cell(row=row, column=5, value=remarks)  # Final Remarks
            ws.cell(row=row, column=6, value=metadata.aadhaar_number or "N/A")  # Aadhaar Number
            ws.cell(row=row, column=7, value=metadata.name or "N/A")  # Name
            ws.cell(row=row, column=8, value=metadata.date_of_birth or "N/A")  # Date of Birth
            ws.cell(row=row, column=9, value=metadata.gender or "N/A")  # Gender
            ws.cell(row=row, column=10, value=metadata.address or "N/A")  # Address
            ws.cell(row=row, column=11, value=f"{metadata.confidence_score*100:.1f}%" if metadata.confidence_score else "N/A")  # Confidence Score
            ws.cell(row=row, column=12, value="Yes" if metadata.is_authentic else "No")  # Is Authentic
            ws.cell(row=row, column=13, value=f"{risk_score:.2f}")  # Risk Score
            ws.cell(row=row, column=14, value=risk_level.title())  # Risk Level
            ws.cell(row=row, column=15, value=yolo_str)  # YOLO Detections
            ws.cell(row=row, column=16, value=cv_str)  # CV Analysis Results
            ws.cell(row=row, column=17, value=fraud_indicators_str)  # Fraud Indicators
            ws.cell(row=row, column=18, value=quality_issues_str)  # Quality Issues
            ws.cell(row=row, column=19, value=doc.uploaded_at.strftime("%Y-%m-%d %H:%M"))  # Upload Date
            ws.cell(row=row, column=20, value=metadata.analyzed_at.strftime("%Y-%m-%d %H:%M") if metadata.analyzed_at else "N/A")  # Analyzed At
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"verification_results_{timestamp}.xlsx"
        
        # Create response
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @action(detail=False, methods=['get'])
    def export_extracted_data(self, request):
        """
        Export extracted data in structured format (CSV/JSON/Excel)
        
        Query params:
        - format: csv, json, or excel (default: csv)
        - document_ids: Optional list of document IDs to export
        
        Returns:
        - Structured data file with extracted information
        """
        from django.db.models import Q
        from django.http import HttpResponse, JsonResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        import csv
        import json
        import datetime
        
        # Get export format
        export_format = request.query_params.get('format', 'csv').lower()
        
        # Get document IDs if specified (filtered by user via get_queryset)
        document_ids = request.query_params.get('document_ids')
        base_queryset = self.get_queryset()
        if document_ids:
            document_ids = [int(id.strip()) for id in document_ids.split(',')]
            documents = (
                base_queryset
                .filter(id__in=document_ids)
                .filter(status='completed')
                .filter(Q(metadata__isnull=False))
                .select_related('metadata')
                .order_by('-uploaded_at')
            )
        else:
            documents = (
                base_queryset
                .filter(status='completed')
                .filter(Q(metadata__isnull=False))
                .select_related('metadata')
                .order_by('-uploaded_at')
            )
        
        # Prepare extracted data
        extracted_data = []
        for doc in documents:
            metadata = doc.metadata
            extracted_data.append({
                'document_id': doc.id,
                'file_name': doc.file_name,
                'upload_date': doc.uploaded_at.strftime("%Y-%m-%d %H:%M:%S"),
                'aadhaar_number': metadata.aadhaar_number or '',
                'name': metadata.name or '',
                'date_of_birth': metadata.date_of_birth or '',
                'gender': metadata.gender or '',
                'address': metadata.address or '',
                'confidence_score': f"{metadata.confidence_score*100:.1f}%" if metadata.confidence_score else '0%',
                'is_authentic': 'Yes' if metadata.is_authentic else 'No',
                'fraud_indicators': '; '.join(metadata.fraud_indicators) if metadata.fraud_indicators else '',
                'quality_issues': '; '.join(metadata.quality_issues) if metadata.quality_issues else '',
                'analyzed_at': metadata.analyzed_at.strftime("%Y-%m-%d %H:%M:%S") if metadata.analyzed_at else ''
            })
        
        # Generate filename with timestamp
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if export_format == 'json':
            filename = f"extracted_data_{timestamp}.json"
            response = JsonResponse(extracted_data, safe=False)
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        elif export_format == 'excel':
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Extracted Data"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Add headers
            headers = ['Document ID', 'File Name', 'Upload Date', 'Aadhaar Number', 'Name', 
                      'Date of Birth', 'Gender', 'Address', 'Confidence Score', 'Is Authentic', 
                      'Fraud Indicators', 'Quality Issues', 'Analyzed At']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add data rows
            for row, data in enumerate(extracted_data, 2):
                ws.cell(row=row, column=1, value=data['document_id'])
                ws.cell(row=row, column=2, value=data['file_name'])
                ws.cell(row=row, column=3, value=data['upload_date'])
                ws.cell(row=row, column=4, value=data['aadhaar_number'])
                ws.cell(row=row, column=5, value=data['name'])
                ws.cell(row=row, column=6, value=data['date_of_birth'])
                ws.cell(row=row, column=7, value=data['gender'])
                ws.cell(row=row, column=8, value=data['address'])
                ws.cell(row=row, column=9, value=data['confidence_score'])
                ws.cell(row=row, column=10, value=data['is_authentic'])
                ws.cell(row=row, column=11, value=data['fraud_indicators'])
                ws.cell(row=row, column=12, value=data['quality_issues'])
                ws.cell(row=row, column=13, value=data['analyzed_at'])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            filename = f"extracted_data_{timestamp}.xlsx"
            response = HttpResponse(
                excel_file.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        else:  # CSV format (default)
            filename = f"extracted_data_{timestamp}.csv"
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            writer = csv.writer(response)
            # Write headers
            writer.writerow(['Document ID', 'File Name', 'Upload Date', 'Aadhaar Number', 'Name', 
                           'Date of Birth', 'Gender', 'Address', 'Confidence Score', 'Is Authentic', 
                           'Fraud Indicators', 'Quality Issues', 'Analyzed At'])
            
            # Write data rows
            for data in extracted_data:
                writer.writerow([
                    data['document_id'],
                    data['file_name'],
                    data['upload_date'],
                    data['aadhaar_number'],
                    data['name'],
                    data['date_of_birth'],
                    data['gender'],
                    data['address'],
                    data['confidence_score'],
                    data['is_authentic'],
                    data['fraud_indicators'],
                    data['quality_issues'],
                    data['analyzed_at']
                ])
            
            return response
    
    @action(detail=False, methods=['get'])
    def fraud_analysis(self, request):
        """
        Get detailed fraud analysis for a document
        
        Query params:
        - document_id: ID of the document to analyze
        - reanalyze: If 'true', re-run fraud detection even if already analyzed
        
        Returns:
        - Detailed fraud detection results with YOLO detections and CV analysis
        """
        from .fraud_detector import detect_fraud
        import json
        
        document_id = request.query_params.get('document_id')
        reanalyze = request.query_params.get('reanalyze', 'false').lower() == 'true'
        
        if not document_id:
            return Response(
                {'error': 'document_id parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Filter by user to ensure user can only view their own documents
            document = self.get_queryset().get(id=document_id)
            metadata, created = DocumentMetadata.objects.get_or_create(document=document)
            
            # Check if fraud detection already exists and reanalyze is not requested
            if not reanalyze and metadata.fraud_detection and 'error' not in metadata.fraud_detection:
                return Response({
                    'document_id': document.id,
                    'file_name': document.file_name,
                    'fraud_detection': metadata.fraud_detection,
                    'status': 'completed'
                })
            
            # Run fraud detection
            image_path = document.preprocessed_file.path if document.preprocessed_file else document.original_file.path
            fraud_result = detect_fraud(image_path)
            
            # Store results
            metadata.fraud_detection = {
                'risk_score': fraud_result.get('risk_score', 0.0),
                'risk_level': fraud_result.get('risk_level', 'low'),
                'yolo_detections': fraud_result.get('detections', []),
                'cv_analysis': fraud_result.get('analysis_details', {}),
                'fraud_indicators': fraud_result.get('fraud_indicators', []),
                'analysis_timestamp': timezone.now().isoformat()
            }
            metadata.save()
            
            return Response({
                'document_id': document.id,
                'file_name': document.file_name,
                'fraud_detection': metadata.fraud_detection,
                'status': 'completed'
            })
            
        except AadhaarDocument.DoesNotExist:
            return Response(
                {'error': 'Document not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            import traceback
            print(f"Fraud analysis error: {traceback.format_exc()}")
            return Response(
                {'error': f'Fraud analysis failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _analyze_document(self, document, metadata):
        """
        Internal method to analyze a document with Gemini and YOLO fraud detection
        
        Args:
            document: AadhaarDocument instance
            metadata: DocumentMetadata instance
        """
        import tempfile
        import os
        
        logger.info(f"Starting analysis for document {document.id}, storage_type={document.storage_type}")
        logger.info(f"Document paths: original_file={document.original_file}, preprocessed_file={document.preprocessed_file}")
        logger.info(f"Supabase paths: original={document.supabase_original_path}, processed={document.supabase_processed_path}")
        
        gemini_service = GeminiService()
        storage_service = get_storage_service()
        
        # Auto-detect storage type based on available paths
        has_supabase_paths = bool(document.supabase_processed_path or document.supabase_original_path)
        has_local_paths = bool(
            (document.preprocessed_file and hasattr(document.preprocessed_file, 'path') and document.preprocessed_file.path) or
            (document.original_file and hasattr(document.original_file, 'path') and document.original_file.path)
        )
        
        logger.info(f"Auto-detect: has_supabase_paths={has_supabase_paths}, has_local_paths={has_local_paths}")
        
        # Get image path - prefer Supabase if available, then try local
        if has_supabase_paths:
            # Download file from Supabase to temp location for analysis
            try:
                # Prefer processed file, fallback to original
                supabase_path = document.supabase_processed_path or document.supabase_original_path
                
                logger.info(f"Downloading from Supabase: {supabase_path}")
                
                # Download file bytes from Supabase
                file_bytes = storage_service.download_file(supabase_path)
                
                # Create temp file for Gemini processing
                suffix = os.path.splitext(supabase_path)[1] or '.jpg'
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp_file:
                    tmp_file.write(file_bytes)
                    image_path = tmp_file.name
                    
                logger.info(f"Downloaded Supabase file to temp: {image_path}")
                
            except Exception as e:
                logger.error(f"Failed to download from Supabase: {e}")
                raise ValueError(f"Cannot access document from Supabase: {e}")
        else:
            # Use local file path
            if document.preprocessed_file and hasattr(document.preprocessed_file, 'path') and document.preprocessed_file.path:
                image_path = document.preprocessed_file.path
            elif document.original_file and hasattr(document.original_file, 'path') and document.original_file.path:
                image_path = document.original_file.path
            else:
                raise ValueError(f"No local file path available for document {document.id}. Storage type: {document.storage_type}")
        
        # Track if we need to cleanup temp file (only if downloaded from Supabase)
        temp_file_path = image_path if has_supabase_paths else None
        
        try:
            # Get analysis from Gemini
            analysis = gemini_service.extract_text_from_image(image_path)
            
            # Update metadata with analysis results
            metadata.full_text = analysis.get('full_text', '')
            
            # Clean Aadhaar number - remove spaces and hyphens to fit in 12-char field
            aadhaar_num = analysis.get('aadhaar_number', '')
            if aadhaar_num:
                aadhaar_num = str(aadhaar_num).replace(' ', '').replace('-', '')[:12]
            metadata.aadhaar_number = aadhaar_num if aadhaar_num else None
            metadata.name = analysis.get('name')
            metadata.date_of_birth = analysis.get('date_of_birth')
            metadata.gender = analysis.get('gender')
            metadata.address = analysis.get('address')
            metadata.gemini_response = analysis.get('raw_gemini_response', '')
            metadata.confidence_score = analysis.get('confidence_score')
            metadata.is_authentic = analysis.get('is_authentic')
            metadata.fraud_indicators = analysis.get('fraud_indicators', [])
            
            # Store design compliance check results
            design_compliance = analysis.get('design_compliance', {})
            
            # Check design compliance - only flag critical missing elements
            if design_compliance:
                # Critical elements that MUST be present in any valid Aadhaar
                critical_elements = [
                    'has_ashoka_emblem', 'has_govt_text', 'has_qr_code', 'has_photo'
                ]
                missing_critical = [elem for elem in critical_elements if design_compliance.get(elem) == False]
                
                # If format_valid is explicitly false, it's suspicious
                format_valid = design_compliance.get('format_valid', True)
                
                if missing_critical:
                    # Add critical design issues to fraud indicators
                    for elem in missing_critical:
                        indicator = f"Missing critical element: {elem.replace('has_', '').replace('_', ' ')}"
                        if indicator not in metadata.fraud_indicators:
                            metadata.fraud_indicators.append(indicator)
                    
                    # Override is_authentic only if multiple critical elements are missing
                    if len(missing_critical) >= 2:
                        metadata.is_authentic = False
                        if 'Document missing critical Aadhaar elements' not in metadata.fraud_indicators:
                            metadata.fraud_indicators.append('Document missing critical Aadhaar elements')
                
                # If format is explicitly invalid
                if format_valid == False:
                    metadata.is_authentic = False
                    if 'Document format does not match any valid Aadhaar format' not in metadata.fraud_indicators:
                        metadata.fraud_indicators.append('Document format does not match any valid Aadhaar format')
            
            # Merge quality issues
            existing_issues = metadata.quality_issues or []
            new_issues = analysis.get('quality_issues', [])
            metadata.quality_issues = list(set(existing_issues + new_issues))
            
            # Store design compliance in extracted_fields for frontend display
            extracted_fields = analysis.get('extracted_fields', {})
            extracted_fields['design_compliance'] = design_compliance
            metadata.extracted_fields = extracted_fields
            
            # Run YOLO fraud detection (async/optional to avoid slowing down main analysis)
            try:
                from .fraud_detector import detect_fraud
                fraud_result = detect_fraud(image_path)
                
                # Store fraud detection results
                metadata.fraud_detection = {
                    'risk_score': fraud_result.get('risk_score', 0.0),
                    'risk_level': fraud_result.get('risk_level', 'low'),
                    'yolo_detections': fraud_result.get('detections', []),
                    'cv_analysis': fraud_result.get('analysis_details', {}),
                    'fraud_indicators': fraud_result.get('fraud_indicators', [])
                }
                
                # Merge fraud indicators from YOLO detection
                # But filter out CV-based false positives (compression, noise, edge artifacts)
                cv_false_positive_keywords = ['compression', 'noise', 'copy-paste', 'edge']
                yolo_indicators = fraud_result.get('fraud_indicators', [])
                
                # Separate critical indicators from CV-based indicators
                critical_indicators = [
                    ind for ind in yolo_indicators
                    if not any(kw in ind.lower() for kw in cv_false_positive_keywords)
                ]
                cv_indicators = [
                    ind for ind in yolo_indicators
                    if any(kw in ind.lower() for kw in cv_false_positive_keywords)
                ]
                
                if yolo_indicators:
                    all_fraud_indicators = metadata.fraud_indicators or []
                    # Add all indicators for display, but CV ones have lower weight
                    all_fraud_indicators.extend(yolo_indicators)
                    metadata.fraud_indicators = list(set(all_fraud_indicators))
                
                # Update authenticity based on fraud detection
                # IMPORTANT: Respect Gemini's verdict unless YOLO finds actual fraud classes
                # CV analysis artifacts (compression, noise) should NOT override Gemini
                fraud_risk_score = fraud_result.get('risk_score', 0.0)
                has_critical_fraud = len(critical_indicators) > 0
                
                # Only override Gemini's "authentic" verdict if:
                # 1. Risk score is very high (>0.7) AND there are critical (non-CV) indicators
                # 2. OR Gemini already said it's not authentic
                if metadata.is_authentic is True:
                    # Gemini said authentic - only override with strong evidence
                    if fraud_risk_score > 0.7 and has_critical_fraud:
                        metadata.is_authentic = False
                        if 'High fraud risk detected by YOLO' not in metadata.fraud_indicators:
                            metadata.fraud_indicators.append('High fraud risk detected by YOLO')
                    # Don't change authenticity just for CV artifacts
                elif metadata.is_authentic is None or metadata.is_authentic is False:
                    # Gemini didn't confirm authentic - use fraud detection result
                    if fraud_risk_score > 0.5 and has_critical_fraud:
                        metadata.is_authentic = False
                        if 'Suspicious document' not in metadata.fraud_indicators:
                            metadata.fraud_indicators.append('Suspicious document')
                    elif fraud_risk_score <= 0.3 and not has_critical_fraud:
                        # Low risk and no critical indicators - might be authentic
                        if metadata.is_authentic is None:
                            metadata.is_authentic = True
                
            except Exception as e:
                # Log error but don't fail the entire analysis
                logger.warning(f"Fraud detection failed for document {document.id}: {e}")
                metadata.fraud_detection = {
                    'error': str(e),
                    'risk_score': 0.0,
                    'risk_level': 'low'
                }
            
            metadata.save()
            
        finally:
            # Cleanup temp file if we created one
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.remove(temp_file_path)
                    logger.info(f"Cleaned up temp file: {temp_file_path}")
                except Exception as e:
                    logger.warning(f"Failed to cleanup temp file: {e}")
